import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def generar_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tablas"

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    fondo_titulo = PatternFill(
        start_color="A9CBE4", end_color="A9CBE4", fill_type="solid"
    )
    fuente_titulo = Font(bold=True, color="000000")

    fondo_fase = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    fondo_neutro = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    fondo_blanco = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )

    fuente_negrita = Font(bold=True)

    fila_actual = 2
    nombre_col_ah = df.columns[33]

    for index, fila in df.iterrows():
        titulo = str(fila.get(nombre_col_ah, f"Registro_{index+1}"))

        ws.merge_cells(
            start_row=fila_actual,
            start_column=1,
            end_row=fila_actual,
            end_column=2,
        )
        celda_titulo = ws.cell(
            row=fila_actual, column=1, value=f"Protección {titulo}"
        )

        celda_titulo.font = fuente_titulo
        celda_titulo.fill = fondo_titulo
        celda_titulo.alignment = Alignment(
            horizontal="center", vertical="center"
        )

        ws.cell(row=fila_actual, column=1).border = borde
        ws.cell(row=fila_actual, column=2).border = borde

        fila_actual += 1

        ct_prim = fila.get("CT_Primario_A", 1)
        ct_sec = fila.get("CT_Secundario_A", 1)
        if pd.isna(ct_sec) or ct_sec == 0:
            ct_sec = 1

        marca_rele = str(fila.get("Marca", "-")).strip().upper()
        es_efacec = marca_rele == "EFACEC"

        def fmt_min_op(val_sec, is_51=False):
            if pd.isna(val_sec) or str(val_sec).strip().upper() in [
                "",
                "-",
                "NO",
                "NAN",
                "NONE",
            ]:
                return "-"
            try:
                val = float(str(val_sec).replace(",", "."))

                if is_51 and es_efacec:
                    val = val * 1.2

                prim = val * (float(ct_prim) / float(ct_sec))

                str_sec = f"{val:g}".replace(".", ",")
                str_prim = f"{prim:g}".replace(".", ",")

                return f"{str_sec} [A-sec] / {str_prim} [A-prim]"
            except ValueError:
                return str(val_sec)

        ctr_str = (
            f"{int(ct_prim)}/{int(ct_sec)}"
            if pd.notna(ct_prim) and pd.notna(ct_sec)
            else fila.get("CTR", "-")
        )

        datos = [
            ("Marca", fila.get("Marca", "-")),
            ("Modelo", fila.get("Modelo", fila.get("Rele_Modelo"))),
        ]

        estado_fase = (
            "Habilitado"
            if str(fila.get("51-1")).strip().upper() == "SI"
            else "Deshabilitado"
        )
        datos.append(("Ajuste de fase", estado_fase))

        if estado_fase == "Habilitado":
            datos.extend(
                [
                    ("CTR", ctr_str),
                    (
                        "Mínimo de Operación",
                        fmt_min_op(fila.get("51-P"), is_51=True),
                    ),
                    ("Lever", fila.get("51-TD")),
                    ("Curva", fila.get("51-C")),
                ]
            )

        val_50_p1 = str(fila.get("50-P1")).strip().upper()

        if val_50_p1 not in ["NO", "NAN", "NONE", "OFF", "", "-"]:
            datos.extend(
                [
                    ("Unidad instantánea", fila.get("50-1")),
                    (
                        "Mínimo de Operación",
                        fmt_min_op(fila.get("50-P1"), is_51=False),
                    ),
                    ("Tiempo de operación", fila.get("50-TD1")),
                ]
            )
            val_50_p2 = str(fila.get("50-P2")).strip().upper()
            if val_50_p2 not in ["NO", "NAN", "NONE", "OFF", "", "-"]:
                datos.extend(
                    [
                        (
                            "Mínimo de Operación",
                            fmt_min_op(fila.get("50-P2"), is_51=False),
                        ),
                        ("Tiempo de operación", fila.get("50-TD2")),
                    ]
                )

        estado_neutro = (
            "Habilitado"
            if str(fila.get("51N-1")).strip().upper() == "SI"
            else "Deshabilitado"
        )
        datos.append(("Ajuste residual", estado_neutro))

        if estado_neutro == "Habilitado":
            datos.extend(
                [
                    ("CTR", ctr_str),
                    (
                        "Mínimo de Operación",
                        fmt_min_op(fila.get("51N-P"), is_51=True),
                    ),
                    ("Lever", fila.get("51N-TD")),
                    ("Curva", fila.get("51N-C")),
                ]
            )

        val_50n_p1 = str(fila.get("50N-P1")).strip().upper()

        if val_50n_p1 not in ["NO", "NAN", "NONE", "OFF", "", "-"]:
            datos.extend(
                [
                    ("Unidad instantánea", fila.get("50N-1")),
                    (
                        "Mínimo de Operación",
                        fmt_min_op(fila.get("50N-P1"), is_51=False),
                    ),
                    ("Tiempo de operación", fila.get("50N-TD1")),
                ]
            )
            val_50n_p2 = str(fila.get("50N-P2")).strip().upper()
            if val_50n_p2 not in ["NO", "NAN", "NONE", "OFF", "", "-"]:
                datos.extend(
                    [
                        (
                            "Mínimo de Operación",
                            fmt_min_op(fila.get("50N-P2"), is_51=False),
                        ),
                        ("Tiempo de operación", fila.get("50N-TD2")),
                    ]
                )

        for param, val in datos:
            valor = str(val) if pd.notna(val) else "-"

            c_param = ws.cell(row=fila_actual, column=1, value=param)
            c_val = ws.cell(row=fila_actual, column=2, value=valor)

            c_param.border = borde
            c_val.border = borde
            c_val.alignment = Alignment(horizontal="right")

            if param == "Ajuste de fase":
                c_param.fill = fondo_fase
                c_val.fill = fondo_fase
                c_param.font = fuente_negrita
            elif param == "Ajuste residual":
                c_param.fill = fondo_neutro
                c_val.fill = fondo_neutro
                c_param.font = fuente_negrita
            else:
                c_param.fill = fondo_blanco
                c_val.fill = fondo_blanco

            fila_actual += 1

        fila_actual += 2

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    # Guardar en memoria en lugar de disco
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- INTERFAZ STREAMLIT ---
st.title("Generador de Tablas hacia Abajo")

archivo_subido = st.file_uploader(
    "Sube el archivo Excel de origen", type=["xlsx"]
)

if archivo_subido:
    try:
        df = pd.read_excel(archivo_subido, sheet_name="Hoja1")

        if st.button("Procesar archivo"):
            excel_buffer = generar_excel(df)

            st.success("Tablas generadas correctamente.")
            st.download_button(
                label="📥 Descargar tablas_verticales.xlsx",
                data=excel_buffer,
                file_name="tablas_verticales.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except Exception as e:
        st.error(f"Ocurrió un error al procesar el archivo: {e}")